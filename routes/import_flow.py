"""Excel 导入流程：上传、预览、异步导入、SSE 进度、停止/取消。

v2 流式优化（2026-07）：
- 上传 / 导入均改用 openpyxl read_only 逐行读取，不再用 pandas 全量载入。
  内存占用恒定在 ~120MB 量级，与文件大小无关（21 万行实测亦稳定）。
- 字段级 diff：只有当工商字段真正发生变化时才 UPDATE 并刷新 updated_at；
  重复但无变化的记录直接跳过，避免时间戳被同源数据重新导入污染。
- 电话 / 股东按归一化值增量合并（仅追加新值、不覆盖主号），命中记录一律执行，
  因此「重复但有新电话」的场景仍会被补全。
- 去重逻辑与文件日期彻底解耦：不再用 file_date 与 updated_at 比大小决定分支。
"""
import os
import re
import json
import io
import uuid
import hashlib
import sqlite3
import tempfile
import threading
import random
import queue as queue_module
from datetime import datetime
from urllib.parse import quote

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from flask import Blueprint, g, request, render_template, redirect, url_for, \
     flash, jsonify, Response

from db import DB_PATH
import backup
import tasks
from queries import invalidate_cache
from data_helpers import (
    split_phones, _split_recommended, split_shareholders,
    split_emails, sync_emails, merge_emails,
)
from utils import (
    map_columns, clean_val, is_industrial_park_file,
    extract_date_from_filename,
    normalize_name, normalize_phone, normalize_credit_code,
    normalize_person_name, normalize_email,
)

bp = Blueprint('import_flow_bp', __name__)


# 全量字段列表（导入流程独立维护，避免与 companies 模块循环依赖）
IMPORT_FIELDS = [
    "name", "phone", "address", "annual_report_address",
    "credit_code", "taxpayer_id", "registration_no", "org_code",
    "legal_person", "registered_capital", "paid_capital",
    "established_date", "approved_date", "business_term",
    "province", "city", "district", "insured_count",
    "company_type", "industry", "former_name", "website",
    "email", "business_scope", "business_status",
    "enterprise_scale", "shareholders", "mailing_address",
    "english_name", "source_file",
]

# 参与字段级 diff（签名比对）的列。
# 注意：不含 source_file（文件名变化不应触发更新），也不含
# phone/email/shareholders（由增量合并单独处理）。
HASH_COLS = [
    "normalized_name", "credit_code", "address", "annual_report_address",
    "mailing_address", "taxpayer_id", "registration_no", "org_code",
    "normalized_legal_person", "registered_capital", "paid_capital",
    "established_date", "approved_date", "business_term",
    "province", "city", "district", "insured_count",
    "company_type", "industry", "former_name", "website",
    "business_scope", "business_status",
    "enterprise_scale", "english_name",
]

PREVIEW_ROWS = 30          # 预览展示行数
COMMIT_EVERY = 1000        # 每 N 行提交一次

HEADER_KEYS = {"公司名称", "企业名称", "统一社会信用代码", "法定代表人",
               "登记状态", "经营状态", "联系电话", "注册资本"}

PHONE_SEP_RE = re.compile(r'[;；,，。、/]')


# ── 入口 ────────────────────────────────────────────────────────────────────

@bp.route("/import", methods=["GET"])
def import_page():
    return render_template("import.html")


# ── 导入模板下载 ────────────────────────────────────────────────────────────

# 模板列定义：(字段名, 表头, 是否必填, 说明, 示例值)
# 表头全部使用 COLUMN_ALIASES 可识别的名称，确保导入时 100% 匹配
TEMPLATE_COLUMNS = [
    ("name",                  "企业名称",          True,  "必填，企业全称",                              "示例科技有限公司"),
    ("legal_person",          "法定代表人",        False, "企业法定代表人姓名",                          "张三"),
    ("phone",                 "联系电话",          False, "多个号码用分号 ; 分隔",                       "0571-88889999;13800138000"),
    ("credit_code",           "统一社会信用代码",  False, "18 位统一社会信用代码",                       "91330100MA12345678"),
    ("taxpayer_id",           "纳税人识别号",      False, "纳税人识别号",                               ""),
    ("registration_no",       "注册号",            False, "工商注册号",                                 ""),
    ("org_code",              "组织机构代码",      False, "组织机构代码",                               ""),
    ("registered_capital",    "注册资本",          False, "如：1000万人民币",                           "1000万人民币"),
    ("paid_capital",          "实缴资本",          False, "如：500万人民币",                            "500万人民币"),
    ("established_date",      "成立日期",          False, "格式 YYYY-MM-DD",                           "2020-01-01"),
    ("approved_date",         "核准日期",          False, "格式 YYYY-MM-DD",                           "2024-06-15"),
    ("business_term",         "营业期限",          False, "如：2020-01-01 至 2050-01-01 或 长期",      "长期"),
    ("business_status",       "经营状态",          False, "存续/注销/吊销/停业等",                      "存续"),
    ("company_type",          "公司类型",          False, "如：有限责任公司",                           "有限责任公司"),
    ("industry",              "所属行业",          False, "国民经济行业分类",                           "软件和信息技术服务业"),
    ("enterprise_scale",      "企业规模",          False, "大型/中型/小型/微型",                         "小型"),
    ("insured_count",         "参保人数",          False, "社保参保人数（整数）",                      "50"),
    ("province",              "省份",              False, "如：浙江省",                                 "浙江省"),
    ("city",                  "所属城市",          False, "如：杭州市",                                 "杭州市"),
    ("district",              "区县",              False, "如：西湖区",                                 "西湖区"),
    ("address",               "注册地址",          False, "工商注册地址",                               "浙江省杭州市西湖区文三路100号"),
    ("annual_report_address", "最新年报地址",      False, "最新年度报告中的地址",                       ""),
    ("mailing_address",       "通信地址",          False, "通信地址",                                  ""),
    ("former_name",           "曾用名",            False, "多个用分号分隔",                             ""),
    ("english_name",          "英文名",            False, "企业英文名称",                               ""),
    ("website",               "网址",              False, "企业官网",                                  "www.example.com"),
    ("email",                 "邮箱",              False, "多个邮箱用分号 ; 分隔",                      "contact@example.com;info@example.com"),
    ("business_scope",        "经营范围",          False, "经营范围全文",                              "技术开发、技术服务、技术咨询"),
    ("shareholders",          "股东",              False, "多个股东用分号 ; 分隔",                      "张三;李四"),
    ("tags",                  "标签",              False, "多个标签用分号 ; 分隔",                      "重点客户;待跟进"),
]

# 列宽配置（按字段名）
TEMPLATE_COL_WIDTHS = {
    "name": 30, "legal_person": 10, "phone": 25,
    "credit_code": 22, "taxpayer_id": 20, "registration_no": 18,
    "org_code": 14, "registered_capital": 14, "paid_capital": 14,
    "established_date": 12, "approved_date": 12, "business_term": 20,
    "business_status": 10, "company_type": 14, "industry": 18,
    "enterprise_scale": 8, "insured_count": 8,
    "province": 6, "city": 8, "district": 10,
    "address": 35, "annual_report_address": 35, "mailing_address": 30,
    "former_name": 18, "english_name": 20, "website": 20,
    "email": 25, "business_scope": 40,
    "shareholders": 20, "tags": 15,
}


@bp.route("/import/template")
def import_template():
    """下载空白导入模板 Excel。

    - Sheet 1「企业数据」：带正确表头 + 一行灰色示例（可删）
    - Sheet 2「字段说明」：字段名、是否必填、格式说明
    所有表头均使用导入可识别的标准名称，填好后直接上传即可。
    """
    wb = openpyxl.Workbook()

    # ── Sheet 1: 企业数据 ──
    ws = wb.active
    ws.title = "企业数据"

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="D97757", end_color="D97757",
                              fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    sample_font = Font(italic=True, size=10, color="999999")
    sample_fill = PatternFill(start_color="FFF8F5", end_color="FFF8F5",
                              fill_type="solid")

    # 写表头
    headers = [col[1] for col in TEMPLATE_COLUMNS]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 写示例行（第 2 行，灰色斜体，用户可删除）
    samples = [col[4] for col in TEMPLATE_COLUMNS]
    for col_idx, val in enumerate(samples, 1):
        cell = ws.cell(row=2, column=col_idx, value=val)
        cell.font = sample_font
        cell.fill = sample_fill
        cell.border = thin_border

    # 列宽
    for col_idx, (key, _hdr, _req, _desc, _sample) in enumerate(TEMPLATE_COLUMNS, 1):
        letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[letter].width = TEMPLATE_COL_WIDTHS.get(key, 15)

    # 冻结首行
    ws.freeze_panes = "A2"

    # ── Sheet 2: 字段说明 ──
    ws2 = wb.create_sheet("字段说明")
    guide_headers = ["字段", "表头名称", "是否必填", "格式说明"]
    guide_font = Font(bold=True, size=11)
    guide_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0",
                             fill_type="solid")
    guide_align = Alignment(horizontal="center", vertical="center")

    for col_idx, h in enumerate(guide_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = guide_font
        cell.fill = guide_fill
        cell.alignment = guide_align
        cell.border = thin_border

    for row_idx, (field, header, required, desc, _sample) in enumerate(TEMPLATE_COLUMNS, 2):
        row_data = [field, header, "必填" if required else "选填", desc]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 8
    ws2.column_dimensions["D"].width = 45
    ws2.freeze_panes = "A2"

    # ── Sheet 3: 使用说明 ──
    ws3 = wb.create_sheet("使用说明")
    instructions = [
        ("导入模板使用说明", True),
        ("", False),
        ("1. 在「企业数据」表中填写企业信息，第一行为表头请勿修改。", False),
        ("2. 第二行为示例数据（灰色斜体），可删除后从第二行开始填真实数据。", False),
        ("3. 「企业名称」为必填项，其余字段选填，留空不会覆盖已有数据。", False),
        ("4. 多值字段（其他电话、股东、标签、其他邮箱、曾用名）用分号 ; 分隔。", False),
        ("5. 表头名称已与导入系统匹配，请勿重命名，否则该列数据将无法导入。", False),
        ("6. 日期格式建议：YYYY-MM-DD（如 2024-01-15）。", False),
        ("7. 填好后保存为 .xlsx 格式，在导入页面上传即可。", False),
        ("", False),
        ("去重规则：", True),
        ("- 优先按统一社会信用代码去重，无信用代码时按企业名称匹配。", False),
        ("- 重复企业只追加新电话，不覆盖已有主号和工商数据。", False),
        ("- 空字段不会清空已有数据，只更新非空且不同的字段。", False),
    ]
    for row_idx, (text, is_heading) in enumerate(instructions, 1):
        cell = ws3.cell(row=row_idx, column=1, value=text)
        if is_heading:
            cell.font = Font(bold=True, size=13)
        else:
            cell.font = Font(size=11)
        cell.alignment = Alignment(vertical="center")
    ws3.column_dimensions["A"].width = 80

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = "EntHub导入模板.xlsx"
    encoded = quote(filename)
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded}",
        }
    )


# ── 表头探测（流式） ────────────────────────────────────────────────────────

def detect_header_row(first_rows):
    """从前若干行中找出真正的表头行索引（0-based）。

    天眼查等导出文件前面常带若干垃圾行；这里只读前几行即可判定，
    不需要把整个文件载入内存。
    """
    for i, row in enumerate(first_rows):
        vals = [str(v).strip() for v in row if v is not None]
        if any(v in HEADER_KEYS for v in vals):
            return i
    return None


# ── 上传（流式分析，不载入全文件） ─────────────────────────────────────────

def _save_upload(file_obj, batch_id, seq):
    """把上传的文件落盘到临时目录，返回路径。"""
    safe = re.sub(r'[^\w.\-]', '_', file_obj.filename or "upload.xlsx")
    path = os.path.join(tempfile.gettempdir(),
                        f"enthub_src_{batch_id}_{seq}_{safe}")
    file_obj.save(path)
    return path


def _meta_path(batch_id):
    return os.path.join(tempfile.gettempdir(), f"enthub_meta_{batch_id}.json")


def _safe_remove(path):
    try:
        if path and os.path.exists(path) and _is_temp_safe(path):
            os.remove(path)
    except Exception:
        pass


def _is_temp_safe(path):
    """路径必须位于系统临时目录内才允许删除。

    防止误删用户原始文件——上传时文件会被复制到临时目录，
    这里只允许删那些临时副本，绝不碰临时目录外的任何文件。
    """
    try:
        abs_path = os.path.abspath(path)
        temp_dir = os.path.abspath(tempfile.gettempdir())
        return abs_path.startswith(temp_dir + os.sep)
    except Exception:
        return False


def _analyze_file(path, filename):
    """用 read_only 模式打开 xlsx，只读表头 + 前若干行，返回分析结果。

    不把整个文件载入内存。返回 dict；若文件不合格抛 ValueError(message)。
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]

        # 读前若干行用于表头探测（顺带取表头）
        # 注意：必须显式指定 max_row，否则 iter_rows 默认用 ws.max_row，
        # 而 read_only 模式下该值可能来自错误的 dimension 元素。
        head_rows = []
        for i, r in enumerate(ws.iter_rows(max_row=6, values_only=True)):
            head_rows.append(r)
            if i >= 4:
                break

        header_idx = detect_header_row(head_rows)
        if header_idx is None:
            header_idx = 0  # 默认第一行

        if header_idx < len(head_rows):
            header = list(head_rows[header_idx])
        else:
            # 表头行不在前 5 行里（极端情况），补读
            header = None
            for r in ws.iter_rows(min_row=header_idx + 1, max_row=header_idx + 1,
                                  values_only=True):
                header = list(r)
                break
            if not header:
                raise ValueError("无法读取表头")

        if is_industrial_park_file(header):
            raise ValueError("检测到非工商数据（工业园区/销售跟踪表），已跳过")

        col_map, sec_phones, rec_phones, unmatched = map_columns(header)
        if "name" not in col_map.values():
            raise ValueError("未找到企业名称列")

        # col_index 必须包含 secondary / recommended 列，否则
        # _row_to_record 里 if sc in col_index / if rc in col_index
        # 永远为 False，更多电话和推荐电话的数据会被静默丢弃。
        col_index = {}
        for orig in list(col_map.keys()) + sec_phones + rec_phones:
            try:
                col_index[orig] = header.index(orig)
            except ValueError:
                pass  # 列名不在 header 中（不应发生，防御性处理）

        # 读前 N 行数据做预览
        preview = []
        data_start = header_idx + 2  # 1-based：表头后第一行
        for i, r in enumerate(ws.iter_rows(
                min_row=data_start, max_row=data_start + PREVIEW_ROWS - 1,
                values_only=True)):
            rec = _row_to_record(r, col_map, col_index, sec_phones, rec_phones,
                                 filename)
            if rec:
                rec["row_num"] = data_start + i
                preview.append(rec)

        # 总行数：ws.max_row 在 read_only 模式下依赖 dimension 元素，
        # 某些 xlsx 文件（天眼查、WPS、旧版 Excel）的 dimension 值是错的，
        # 导致 max_row 返回 3 而实际有 5000+ 行。
        # iter_rows 不带 max_row 时也默认用 ws.max_row，所以必须显式传大值。
        EXCEL_MAX_ROWS = 1048576
        total = ws.max_row or 0
        total = max(0, total - (header_idx + 1))
        if total < len(preview):
            total = sum(1 for _ in ws.iter_rows(
                min_row=header_idx + 2, max_row=EXCEL_MAX_ROWS,
                values_only=True))

        return {
            "header_idx": header_idx,
            "col_map": {str(k): v for k, v in col_map.items()},
            "secondary": [str(c) for c in sec_phones],
            "recommended": [str(c) for c in rec_phones],
            "unmatched": [str(c) for c in unmatched],
            "total_cols": len(header),
            "preview": preview,
            "total": total,
        }
    finally:
        wb.close()


def _row_to_record(row, col_map, col_index, sec_phones, rec_phones, source_name):
    """把一行原始数据清洗成 record dict（与原 pandas 版口径一致）。"""
    if row is None:
        return None
    record = {}
    for orig_col, field in col_map.items():
        idx = col_index[orig_col]
        val = row[idx] if idx < len(row) else None
        record[field] = clean_val(val, field)

    # 副电话列（启信宝：联系电话2~10）合并到 phone
    sec_parts = []
    for sc in sec_phones:
        if sc in col_index:
            idx = col_index[sc]
            v = clean_val(row[idx] if idx < len(row) else None)
            if v:
                sec_parts.append(v)
    if sec_parts:
        existing_phone = record.get("phone", "")
        merged = ";".join(p for p in [existing_phone] + sec_parts if p)
        record["phone"] = merged

    # 推荐电话
    rec_parts = []
    for rc in rec_phones:
        if rc in col_index:
            idx = col_index[rc]
            v = clean_val(row[idx] if idx < len(row) else None)
            if v:
                rec_parts.append(v)
    record["_recommended_phone"] = ";".join(rec_parts)

    if not record.get("source_file"):
        record["source_file"] = source_name

    # 归一化
    record["normalized_name"] = normalize_name(record.get("name", ""))
    record["normalized_phone"] = normalize_phone(record.get("phone", ""))
    if record.get("credit_code"):
        record["credit_code"] = normalize_credit_code(record["credit_code"])

    # 跳过无效占位
    placeholder_texts = ["暂不予显示", "企业信息暂不"]
    for fld in ("name", "business_scope", "address"):
        val = record.get(fld, "")
        if val and any(pt in val for pt in placeholder_texts):
            return None

    if not record.get("name"):
        return None
    return record


@bp.route("/import/upload", methods=["POST"])
def import_upload():
    files = request.files.getlist("file")
    valid_files = [f for f in files if f.filename]

    if not valid_files:
        flash("未选择文件", "error")
        return redirect(url_for("import_flow_bp.import_page"))

    batch_id = uuid.uuid4().hex[:12]
    errors = []
    file_metas = []   # worker 用：path / header_idx / total / file_date
    mappings = []     # 预览页用：展示匹配情况
    all_preview = []  # 预览样本

    for seq, file in enumerate(valid_files):
        path = _save_upload(file, batch_id, seq)
        try:
            info = _analyze_file(path, file.filename)
        except ValueError as e:
            errors.append(f"{file.filename}: {e}")
            _safe_remove(path)
            continue
        except Exception as e:
            errors.append(f"{file.filename}: 读取失败 ({e})")
            _safe_remove(path)
            continue

        file_date = extract_date_from_filename(file.filename)

        file_metas.append({
            "path": path,
            "filename": file.filename,
            "header_idx": info["header_idx"],
            "total": info["total"],
            "file_date": file_date,
        })
        mappings.append({
            "file": file.filename,
            "total_cols": info["total_cols"],
            "matched": info["col_map"],
            "secondary": info["secondary"],
            "recommended": info["recommended"],
            "unmatched": info["unmatched"],
        })
        all_preview.extend(info["preview"])

    if not file_metas:
        for e in errors:
            flash(e, "error")
        return redirect(url_for("import_flow_bp.import_page"))

    total = sum(f["total"] for f in file_metas)

    # 写预览样本到 import_preview 表（仅前 30 条）
    for rec in all_preview[:PREVIEW_ROWS]:
        g.db.execute("""
            INSERT INTO import_preview
                (batch_id, row_num, name, normalized_name, phone, normalized_phone,
                 address, credit_code, legal_person, is_duplicate, duplicate_reason, will_update)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, [
            batch_id, rec.get("row_num", 0), rec.get("name", ""),
            rec.get("normalized_name", ""), rec.get("phone", ""),
            rec.get("normalized_phone", ""), rec.get("address", ""),
            rec.get("credit_code", ""), rec.get("legal_person", ""),
            0, "", 0
        ])
    g.db.commit()

    # 小 metadata 文件（只含路径 / 映射摘要，不含全量数据）
    meta_path = _meta_path(batch_id)
    with open(meta_path, "w") as f:
        json.dump({
            "files": file_metas,
            "mappings": mappings,
            "total": total,
        }, f, ensure_ascii=False)

    if errors:
        for e in errors:
            flash(e, "warning")

    flash(f"已分析 {len(file_metas)} 个文件，共 {total} 条记录", "success")
    return redirect(url_for("import_flow_bp.import_preview", batch_id=batch_id))


# ── 预览 ────────────────────────────────────────────────────────────────────

@bp.route("/import/preview/<batch_id>")
def import_preview(batch_id):
    meta_path = _meta_path(batch_id)
    if not os.path.exists(meta_path):
        flash("导入会话已过期，请重新上传", "error")
        return redirect(url_for("import_flow_bp.import_page"))

    with open(meta_path) as f:
        data = json.load(f)

    mappings = data.get("mappings", [])
    total = data.get("total", 0)

    sample = g.db.execute("""
        SELECT * FROM import_preview
        WHERE batch_id = ? ORDER BY row_num LIMIT ?
    """, [batch_id, PREVIEW_ROWS]).fetchall()

    # 数据质量提示：电话字段含分隔符（多号码塞一格，将自动拆分）
    quality_samples = []
    for row in sample:
        phone_val = row["phone"] if "phone" in row.keys() else None
        if not phone_val:
            continue
        phone_str = str(phone_val).strip()
        if PHONE_SEP_RE.search(phone_str):
            masked = phone_str[:7] + "***" if len(phone_str) > 7 else phone_str
            quality_samples.append(masked)

    return render_template("import_preview.html", batch_id=batch_id,
                           total=total, sample=sample, mappings=mappings,
                           quality_samples=quality_samples)


# ── 确认导入 ────────────────────────────────────────────────────────────────

@bp.route("/import/confirm/<batch_id>", methods=["POST"])
def import_confirm(batch_id):
    skip_dup = request.form.get("skip_dup", "1") == "1"

    meta_path = _meta_path(batch_id)
    if not os.path.exists(meta_path):
        flash("导入会话已过期，请重新上传", "error")
        return redirect(url_for("import_flow_bp.import_page"))

    with open(meta_path) as f:
        meta = json.load(f)

    # 导入前自动备份
    backup_result = backup.create_backup(DB_PATH, reason="导入前自动备份")
    if backup_result["success"]:
        backup.cleanup_old_backups(keep_count=7)
        backup_msg = f"已自动创建备份：{backup_result['filename']}"
    else:
        backup_msg = f"自动备份失败：{backup_result.get('error', '未知错误')}"

    task_queue, stop_event = tasks.create(batch_id)
    t = threading.Thread(
        target=_import_worker,
        args=(batch_id, meta, skip_dup, task_queue, stop_event),
        daemon=True,
    )
    t.start()

    return render_template("import_progress.html",
                           batch_id=batch_id, backup_msg=backup_msg)


# ── 字段签名（diff 判定） ───────────────────────────────────────────────────

def _signature(values):
    """对一组字段值计算 md5 签名，用于判断工商字段是否发生变化。"""
    joined = "\x1f".join("" if v is None else str(v) for v in values)
    return hashlib.md5(joined.encode("utf-8")).hexdigest()


def _incoming_values(rec):
    """从导入 record 计算 HASH_COLS 对齐的字段值列表（归一化后）。"""
    return [
        rec.get("normalized_name", ""),
        rec.get("credit_code", ""),
        rec.get("address", ""),
        rec.get("annual_report_address", ""),
        rec.get("mailing_address", ""),
        rec.get("taxpayer_id", ""),
        rec.get("registration_no", ""),
        rec.get("org_code", ""),
        normalize_person_name(rec.get("legal_person", "")),
        rec.get("registered_capital", ""),
        rec.get("paid_capital", ""),
        rec.get("established_date", ""),
        rec.get("approved_date", ""),
        rec.get("business_term", ""),
        rec.get("province", ""),
        rec.get("city", ""),
        rec.get("district", ""),
        rec.get("insured_count", ""),
        rec.get("company_type", ""),
        rec.get("industry", ""),
        rec.get("former_name", ""),
        rec.get("website", ""),
        rec.get("business_scope", ""),
        rec.get("business_status", ""),
        rec.get("enterprise_scale", ""),
        rec.get("english_name", ""),
    ]


def _incoming_signature(rec):
    """从导入 record 计算签名（与 DB 中存储的签名口径一致）。"""
    return _signature(_incoming_values(rec))


def _has_real_changes(db, company_id, rec, hash_cols_sql):
    """签名不同时的精确判定：是否存在「非空且与库内不同」的传入字段。

    空的传入字段不算变化——文件没带这个字段 ≠ 数据被改成空。
    只有传入值非空、且与库里现存值不同，才算真实变化。
    """
    row = db.execute(
        f"SELECT {hash_cols_sql} FROM companies WHERE id = ?",
        [company_id]
    ).fetchone()
    if not row:
        return True
    for col, inc in zip(HASH_COLS, _incoming_values(rec)):
        if inc and inc != (row[col] or ""):
            return True
    return False


# ── 增量合并（带预加载缓存，避免逐行 SELECT） ──────────────────────────────

def _count_raw_phones(phone_str):
    """Count total phone parts in input string before validation."""
    if not phone_str:
        return 0
    SEPS = [';', '；', ',', '，', '。', '、', '/']
    result = [str(phone_str)]
    for sep in SEPS:
        new_result = []
        for part in result:
            new_result.extend(part.split(sep))
        result = new_result
    return len([p for p in result if p.strip()])


def _merge_phones_cached(db, company_id, phone_str,
                         recommended_str, phone_sets, has_primary):
    """用预加载的号码集合做增量合并，返回新增号码条数。

    语义与 data_helpers.merge_phones 一致：只追加归一化后不重复的新号码，
    不覆盖已有主号；仅当公司无主号时才按优先级设主号（推荐>联系）。
    """
    sset = phone_sets.get(company_id)
    if sset is None:
        sset = set()
        phone_sets[company_id] = sset

    raw_count = _count_raw_phones(phone_str)
    valid_phones = split_phones(phone_str)
    recommended = _split_recommended(recommended_str)
    skipped_invalid = raw_count - len(valid_phones)

    # 确定主号候选（仅当公司无主号时才需要）
    needs_primary = company_id not in has_primary
    primary_norm = None
    if needs_primary:
        if recommended:
            primary_norm = recommended[0][1]
        elif valid_phones:
            primary_norm = valid_phones[0][1]

    added = 0
    for raw, norm in valid_phones:
        if norm and norm not in sset:
            is_primary = 1 if (needs_primary and norm == primary_norm) else 0
            db.execute(
                "INSERT INTO company_phones "
                "(company_id, phone, normalized_phone, is_primary) "
                "VALUES (?, ?, ?, ?)",
                [company_id, raw, norm, is_primary]
            )
            sset.add(norm)
            if is_primary:
                has_primary.add(company_id)
                needs_primary = False
            added += 1

    for raw, norm in recommended:
        if norm and norm not in sset:
            is_primary = 1 if (needs_primary and norm == primary_norm) else 0
            db.execute(
                "INSERT INTO company_phones "
                "(company_id, phone, normalized_phone, is_primary) "
                "VALUES (?, ?, ?, ?)",
                [company_id, raw, norm, is_primary]
            )
            sset.add(norm)
            if is_primary:
                has_primary.add(company_id)
                needs_primary = False
            added += 1

    return added, skipped_invalid


def _merge_shareholders_cached(db, company_id, shareholders_str, sh_sets):
    """用预加载集合做股东增量合并，返回新增条数。"""
    if not shareholders_str:
        return 0
    sset = sh_sets.get(company_id)
    if sset is None:
        sset = set()
        sh_sets[company_id] = sset
    added = 0
    for raw, norm, position in split_shareholders(shareholders_str):
        if norm and norm not in sset:
            db.execute(
                "INSERT INTO company_shareholders "
                "(company_id, name, normalized_name, position) VALUES (?, ?, ?, ?)",
                [company_id, raw, norm, position]
            )
            sset.add(norm)
            added += 1
    return added


def _merge_emails_cached(db, company_id, email_str, email_sets, has_email_primary):
    """用预加载的邮箱集合做增量合并，返回新增邮箱条数。

    语义与 data_helpers.merge_emails 一致：只追加归一化后不重复的新邮箱，
    不覆盖已有主邮箱；仅当公司无主邮箱时才把首个新邮箱设为主邮箱。
    """
    if not email_str:
        return 0
    eset = email_sets.get(company_id)
    if eset is None:
        eset = set()
        email_sets[company_id] = eset

    added = 0
    for raw, norm in split_emails(email_str):
        if norm and norm not in eset:
            is_primary = 1 if company_id not in has_email_primary else 0
            db.execute(
                "INSERT INTO company_emails "
                "(company_id, email, normalized_email, is_primary) "
                "VALUES (?, ?, ?, ?)",
                [company_id, raw, norm, is_primary]
            )
            eset.add(norm)
            if is_primary:
                has_email_primary.add(company_id)
            added += 1
    return added


# ── 后台导入线程（流式） ────────────────────────────────────────────────────

def _stream_records(meta_file):
    """生成器：逐行流式读取一个文件，产出清洗后的 record dict。"""
    path = meta_file["path"]
    header_idx = meta_file["header_idx"]
    source_name = meta_file["filename"]

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]

        # 读取表头行（1-based: header_idx+1）
        header = None
        for r in ws.iter_rows(min_row=header_idx + 1, max_row=header_idx + 1,
                              values_only=True):
            header = list(r)
            break
        if not header:
            return

        col_map, sec_phones, rec_phones, _ = map_columns(header)
        # 同 _analyze_file：col_index 必须包含 secondary / recommended 列
        col_index = {}
        for orig in list(col_map.keys()) + sec_phones + rec_phones:
            try:
                col_index[orig] = header.index(orig)
            except ValueError:
                pass

        data_start = header_idx + 2
        for row in ws.iter_rows(min_row=data_start, max_row=1048576,
                                values_only=True):
            rec = _row_to_record(row, col_map, col_index, sec_phones, rec_phones,
                                 source_name)
            if rec:
                yield rec
    finally:
        wb.close()


def _stats(processed, total, inserted, updated, phones_merged, skipped, phones_invalid=0):
    return {
        "processed": processed, "total": total,
        "inserted": inserted, "updated": updated,
        "phones_merged": phones_merged, "skipped": skipped,
        "phones_invalid": phones_invalid,
    }


# 预设色板（自动创建标签时随机选色）
_TAG_COLORS = [
    "#ef4444", "#f97316", "#eab308", "#22c55e", "#06b6d4",
    "#3b82f6", "#8b5cf6", "#ec4899", "#6b7280",
]

# 标签分隔符（中英文分号）
_TAG_SEP_RE = re.compile(r'[;；]')


def _sync_tags(db, company_id, tags_str):
    """将分号分隔的标签文本转为结构化标签（tags + company_tags 表）。

    - 按分号拆分标签名
    - tags 表中不存在的标签自动创建（随机颜色）
    - 通过 company_tags 关联（INSERT OR IGNORE 去重，不覆盖已有标签）
    """
    if not tags_str:
        return
    names = [n.strip() for n in _TAG_SEP_RE.split(tags_str) if n.strip()]
    for name in names:
        row = db.execute(
            "SELECT id FROM tags WHERE name = ?", [name]
        ).fetchone()
        if row:
            tag_id = row["id"]
        else:
            color = random.choice(_TAG_COLORS)
            cursor = db.execute(
                "INSERT INTO tags (name, color) VALUES (?, ?)",
                [name, color]
            )
            tag_id = cursor.lastrowid
        db.execute(
            "INSERT OR IGNORE INTO company_tags (company_id, tag_id) VALUES (?, ?)",
            [company_id, tag_id]
        )


def _do_insert(db, rec):
    """构造并执行 INSERT，返回新 id。电话 / 股东 / 标签由调用方单独写入。"""
    fields = {}
    for f_name in IMPORT_FIELDS:
        val = rec.get(f_name, "")
        if val:
            fields[f_name] = val
    fields["normalized_name"] = rec.get("normalized_name", "")
    lp = fields.get("legal_person", "")
    fields["normalized_legal_person"] = normalize_person_name(lp) if lp else ""
    # 这些由调用方单独处理，不写入 companies 主表
    fields.pop("phone", None)
    fields.pop("email", None)
    fields.pop("shareholders", None)
    fields.pop("tags", None)
    fields["status"] = "active"
    fields["source"] = "import"
    # created_at / updated_at 走 DB 默认值（now）
    cols = ", ".join(fields.keys())
    placeholders = ", ".join(["?"] * len(fields))
    cursor = db.execute(
        f"INSERT INTO companies ({cols}) VALUES ({placeholders})",
        list(fields.values())
    )
    return cursor.lastrowid


def _do_update(db, company_id, rec):
    """字段级 UPDATE：写入所有非空工商字段 + 派生归一化字段 + 刷新 updated_at。"""
    fields = {}
    for f_name in IMPORT_FIELDS:
        val = rec.get(f_name, "")
        if val:
            fields[f_name] = val
    fields["normalized_name"] = rec.get("normalized_name", "")
    lp = fields.get("legal_person", "")
    fields["normalized_legal_person"] = normalize_person_name(lp) if lp else ""
    fields.pop("phone", None)
    fields.pop("email", None)
    fields.pop("shareholders", None)
    fields.pop("tags", None)

    fields["updated_at"] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    set_clause = ", ".join([f"{k} = ?" for k in fields.keys()])
    db.execute(
        f"UPDATE companies SET {set_clause} WHERE id = ?",
        list(fields.values()) + [company_id]
    )


def _import_worker(batch_id, meta, skip_dup, task_queue, stop_event):
    """后台流式导入线程：逐条处理，发送进度事件。"""
    def send(event, data=None):
        task_queue.put({"event": event, "data": data or {}})

    db = None
    try:
        files = meta["files"]
        total = meta["total"]
        send("start", {"total": total})

        db = sqlite3.connect(DB_PATH)
        db.row_factory = sqlite3.Row
        db.execute("PRAGMA journal_mode=WAL")

        # ── 预加载索引（内存占用 ~120MB，与文件大小无关）──
        code_index = {}    # credit_code -> id
        name_index = {}    # normalized_name -> id
        sig_index = {}     # id -> 工商字段签名
        phone_sets = {}    # id -> set(normalized_phone)
        has_primary = set()
        sh_sets = {}       # id -> set(normalized_shareholder_name)

        hash_cols_sql = ", ".join(HASH_COLS)
        for row in db.execute(
            f"SELECT id, credit_code, normalized_name, {hash_cols_sql} FROM companies"
        ):
            cid = row["id"]
            if row["credit_code"]:
                code_index[row["credit_code"]] = cid
            if row["normalized_name"]:
                name_index[row["normalized_name"]] = cid
            sig_index[cid] = _signature([row[c] or "" for c in HASH_COLS])

        for row in db.execute(
            "SELECT company_id, normalized_phone, is_primary FROM company_phones"
        ):
            cid = row["company_id"]
            norm = row["normalized_phone"] or ""
            if norm:
                phone_sets.setdefault(cid, set()).add(norm)
            if row["is_primary"]:
                has_primary.add(cid)

        for row in db.execute(
            "SELECT company_id, normalized_name FROM company_shareholders"
        ):
            cid = row["company_id"]
            norm = row["normalized_name"] or ""
            if norm:
                sh_sets.setdefault(cid, set()).add(norm)

        # 预加载邮箱索引
        email_sets = {}    # id -> set(normalized_email)
        has_email_primary = set()
        for row in db.execute(
            "SELECT company_id, normalized_email, is_primary FROM company_emails"
        ):
            cid = row["company_id"]
            norm = row["normalized_email"] or ""
            if norm:
                email_sets.setdefault(cid, set()).add(norm)
            if row["is_primary"]:
                has_email_primary.add(cid)

        # ── 逐文件、逐行处理 ──
        inserted = updated = phones_merged = skipped = phones_invalid = processed = 0
        report_every = max(1, total // 100) if total else 1
        today_ts = None  # 懒加载，仅在有变更时取一次

        for fmeta in files:
            for rec in _stream_records(fmeta):
                if stop_event.is_set():
                    db.commit()
                    send("stopped", _stats(processed, total, inserted, updated,
                                           phones_merged, skipped, phones_invalid))
                    _cleanup(meta, batch_id, db)
                    return

                processed += 1

                # 去重判定：先信用代码，再归一化名称
                existing_id = None
                cc = rec.get("credit_code", "")
                nn = rec.get("normalized_name", "")
                if cc and cc in code_index:
                    existing_id = code_index[cc]
                elif nn and nn in name_index:
                    existing_id = name_index[nn]

                if existing_id is None:
                    # ── 新增 ──
                    new_id = _do_insert(db, rec)
                    if cc:
                        code_index[cc] = new_id
                    if nn:
                        name_index[nn] = new_id
                    sig_index[new_id] = _incoming_signature(rec)

                    # 同步写入电话，并更新缓存
                    pset = set()
                    _raw = rec.get("phone", "")
                    _valid = split_phones(_raw)
                    _recommended = _split_recommended(rec.get("_recommended_phone", ""))
                    phones_invalid += _count_raw_phones(_raw) - len(_valid)

                    # 主号优先级：推荐电话 > 联系电话（第一号）
                    primary_norm = None
                    if _recommended:
                        primary_norm = _recommended[0][1]
                    elif _valid:
                        primary_norm = _valid[0][1]

                    for raw, norm in _valid:
                        is_primary = 1 if norm == primary_norm else 0
                        db.execute(
                            "INSERT INTO company_phones "
                            "(company_id, phone, normalized_phone, is_primary) "
                            "VALUES (?, ?, ?, ?)",
                            [new_id, raw, norm, is_primary]
                        )
                        if norm:
                            pset.add(norm)
                            if is_primary:
                                has_primary.add(new_id)
                    for raw, norm in _recommended:
                        if norm and norm not in pset:
                            is_primary = 1 if norm == primary_norm else 0
                            db.execute(
                                "INSERT INTO company_phones "
                                "(company_id, phone, normalized_phone, "
                                "is_primary) "
                                "VALUES (?, ?, ?, ?)",
                                [new_id, raw, norm, is_primary]
                            )
                            pset.add(norm)
                            if is_primary:
                                has_primary.add(new_id)
                    phone_sets[new_id] = pset
                    # 同步写入邮箱
                    eset = set()
                    for i, (raw, norm) in enumerate(split_emails(rec.get("email", ""))):
                        db.execute(
                            "INSERT INTO company_emails "
                            "(company_id, email, normalized_email, is_primary) "
                            "VALUES (?, ?, ?, ?)",
                            [new_id, raw, norm, 1 if i == 0 else 0]
                        )
                        if norm:
                            eset.add(norm)
                    email_sets[new_id] = eset
                    if eset:
                        has_email_primary.add(new_id)
                    _merge_shareholders_cached(db, new_id,
                                               rec.get("shareholders", ""), sh_sets)
                    _sync_tags(db, new_id, rec.get("tags", ""))
                    inserted += 1
                else:
                    # ── 命中已有：先合并电话（补全），再判断字段是否变化 ──
                    incoming_sig = _incoming_signature(rec)
                    sig_match = (sig_index.get(existing_id) == incoming_sig)

                    _raw = rec.get("phone", "")
                    phone_added, _skipped = _merge_phones_cached(
                        db, existing_id, _raw,
                        rec.get("_recommended_phone", ""),
                        phone_sets, has_primary)
                    phones_invalid += _skipped
                    _sync_tags(db, existing_id, rec.get("tags", ""))

                    if sig_match:
                        real_change = False
                    else:
                        # 签名不同：可能是空字段导致，精确比对非空字段
                        real_change = _has_real_changes(
                            db, existing_id, rec, hash_cols_sql)

                    if real_change and not skip_dup:
                        # 有真实非空字段变化且允许覆盖 → 字段级 UPDATE
                        _do_update(db, existing_id, rec)
                        sig_index[existing_id] = incoming_sig
                        updated += 1
                    elif phone_added:
                        # 仅补了电话 → 数据已变，刷新 updated_at
                        if today_ts is None:
                            today_ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        db.execute(
                            "UPDATE companies SET updated_at = ? WHERE id = ?",
                            [today_ts, existing_id])
                        phones_merged += 1
                    else:
                        # 完全无变化 → 跳过，不动 updated_at
                        skipped += 1

                # 批量提交 + 进度报告
                if processed % COMMIT_EVERY == 0:
                    db.commit()
                if processed % report_every == 0:
                    send("progress", _stats(processed, total, inserted, updated,
                                            phones_merged, skipped, phones_invalid))

        db.commit()
        db.execute("DELETE FROM import_preview WHERE batch_id = ?", [batch_id])
        db.commit()

        # 数据已变更，清空筛选器缓存
        invalidate_cache()
        send("done", _stats(total, total, inserted, updated,
                           phones_merged, skipped, phones_invalid))
    except Exception as e:
        try:
            if db:
                db.commit()
        except Exception:
            pass
        send("error", {"message": str(e)})
    finally:
        _cleanup(meta, batch_id, db)


def _cleanup(meta, batch_id, db):
    """清理临时文件（源 xlsx / metadata）并关闭连接。"""
    try:
        for fmeta in meta.get("files", []):
            _safe_remove(fmeta.get("path"))
        _safe_remove(_meta_path(batch_id))
    except Exception:
        pass
    try:
        if db:
            db.close()
    except Exception:
        pass


# ── SSE 进度推送 ────────────────────────────────────────────────────────────

@bp.route("/import/confirm/<batch_id>/stream")
def import_stream(batch_id):
    """SSE 端点：向前端推送导入进度。"""
    task = tasks.get(batch_id)
    if not task:
        return Response('event: error\ndata: {"message": "任务不存在"}\n\n',
                        content_type="text/event-stream")

    def generate():
        try:
            while True:
                try:
                    msg = task["queue"].get(timeout=30)
                    event = msg.get("event", "message")
                    data = json.dumps(msg.get("data", {}), ensure_ascii=False)
                    yield f"event: {event}\ndata: {data}\n\n"
                    if event in ("done", "error", "stopped"):
                        break
                except queue_module.Empty:
                    yield 'event: heartbeat\ndata: {}\n\n'
        finally:
            tasks.pop(batch_id, None)

    return Response(generate(), content_type="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


# ── 停止 / 取消 ─────────────────────────────────────────────────────────────

@bp.route("/import/confirm/<batch_id>/stop", methods=["POST"])
def import_stop(batch_id):
    if tasks.request_stop(batch_id):
        return jsonify({"ok": True})
    return jsonify({"ok": False}), 404


@bp.route("/import/cancel/<batch_id>", methods=["POST"])
def import_cancel(batch_id):
    meta_path = _meta_path(batch_id)
    if os.path.exists(meta_path):
        try:
            with open(meta_path) as f:
                meta = json.load(f)
            for fmeta in meta.get("files", []):
                _safe_remove(fmeta.get("path"))
        except Exception:
            pass
        _safe_remove(meta_path)
    flash("已取消导入", "info")
    return redirect(url_for("import_flow_bp.import_page"))
